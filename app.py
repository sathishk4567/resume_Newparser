import os
import re
import io
import json
import sqlite3
from datetime import datetime
from functools import wraps

from flask import (
    Flask, request, render_template, redirect, url_for,
    session, flash, send_file, jsonify
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

import pdfplumber
import docx
import openpyxl
from openpyxl.utils import get_column_letter
import requests

APP_ROOT = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(APP_ROOT, "data")
DB_PATH = os.path.join(DATA_DIR, "app.db")
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024  # 25MB per request

ALLOWED_EXT = {".pdf", ".docx"}
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")


# ---------------------------------------------------------------- database
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS candidates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filename TEXT,
        name TEXT,
        email TEXT,
        phone TEXT,
        linkedin TEXT,
        location TEXT,
        current_title TEXT,
        years_experience TEXT,
        education TEXT,
        skills TEXT,
        summary TEXT,
        visa_status TEXT,
        uploaded_by TEXT,
        uploaded_at TEXT
    );

    CREATE TABLE IF NOT EXISTS experience (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        candidate_id INTEGER NOT NULL,
        title TEXT,
        company TEXT,
        dates TEXT,
        location TEXT,
        highlights TEXT,
        FOREIGN KEY(candidate_id) REFERENCES candidates(id)
    );

    CREATE TABLE IF NOT EXISTS requirements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        label TEXT,
        raw_text TEXT,
        title TEXT,
        location TEXT,
        skills TEXT,
        visa_status TEXT,
        created_by TEXT,
        created_at TEXT
    );
    """)
    conn.commit()

    # migration: add visa_status to candidates table if this DB predates the column
    existing_cols = [r["name"] for r in conn.execute("PRAGMA table_info(candidates)").fetchall()]
    if "visa_status" not in existing_cols:
        conn.execute("ALTER TABLE candidates ADD COLUMN visa_status TEXT")
        conn.commit()

    existing = conn.execute("SELECT COUNT(*) as c FROM users").fetchone()["c"]
    if existing == 0:
        admin_user = os.environ.get("ADMIN_USERNAME", "admin")
        admin_pass = os.environ.get("ADMIN_PASSWORD", "changeme123")
        conn.execute(
            "INSERT INTO users (username, password_hash, created_at) VALUES (?, ?, ?)",
            (admin_user, generate_password_hash(admin_pass), datetime.utcnow().isoformat())
        )
        conn.commit()
    conn.close()


# ---------------------------------------------------------------- auth
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        conn.close()
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            return redirect(url_for("dashboard"))
        flash("Incorrect username or password.")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/users/add", methods=["POST"])
@login_required
def add_user():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    if not username or not password:
        flash("Username and password are required.")
        return redirect(url_for("dashboard"))
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO users (username, password_hash, created_at) VALUES (?, ?, ?)",
            (username, generate_password_hash(password), datetime.utcnow().isoformat())
        )
        conn.commit()
        flash(f"Added teammate '{username}'.")
    except sqlite3.IntegrityError:
        flash("That username already exists.")
    conn.close()
    return redirect(url_for("dashboard"))


# ---------------------------------------------------------------- parsing
SKILLS_DICT = [
    "Python", "Java", "JavaScript", "TypeScript", r"C\+\+", "C#", "Go", "Rust", "Ruby", "PHP",
    "Swift", "Kotlin", "Bash", "SQL", "R", "Scala",
    "FastAPI", "Flask", "Django", "React", "Angular", "Vue", r"Node\.js", "Express", "REST",
    "GraphQL", "OpenAPI", "Pydantic", "Spring", r"\.NET",
    "AWS", "Azure", "GCP", "Google Cloud", "Docker", "Kubernetes", "Terraform", "Jenkins",
    "GitHub Actions", "CI/CD", "Git",
    "PostgreSQL", "MySQL", "MongoDB", "Redis", "DynamoDB", "SQL Server", "Oracle", "Snowflake", "BigQuery",
    "Machine Learning", "Deep Learning", "TensorFlow", "PyTorch", "scikit-learn", "pandas",
    "NumPy", "NLP", "RAG", "LLM", "OpenAI",
    "Excel", "Power BI", "Tableau", "Looker", "Salesforce", "SAP", "Jira", "Confluence",
    "Agile", "Scrum", "Kanban",
    "Project Management", "Product Management", "Leadership", "Communication", "Negotiation",
    "Public Speaking", "Budgeting",
    "Sales", "Marketing", "SEO", "Content Writing", "Copywriting", "Social Media", "Photoshop",
    "Illustrator", "Figma", "AutoCAD",
    "Customer Service", "Recruiting", "HR", "Accounting", "Bookkeeping", "Nursing", "Teaching",
    "Data Analysis", "Data Engineering", "ETL", "A/B Test", "Statistics", "Pytest", "JUnit",
    "Selenium", "QA Testing"
]
TITLE_KEYWORDS = [
    "Engineer", "Manager", "Developer", "Analyst", "Director", "Designer", "Specialist",
    "Consultant", "Lead", "Architect", "Coordinator", "Intern", "Scientist", "Administrator",
    "Executive", "Officer", "Founder", "President", "Recruiter", "Accountant", "Nurse",
    "Teacher", "Marketer", "Representative"
]

# Ordered so more specific phrases are checked before generic ones
VISA_PATTERNS = [
    (r"no\s+sponsorship\s+(?:required|needed)", "No sponsorship required"),
    (r"(?:requires?|needs?)\s+(?:visa\s+)?sponsorship", "Sponsorship required"),
    (r"authorized\s+to\s+work.{0,40}without\s+sponsorship", "Authorized to work (no sponsorship)"),
    (r"u\.?s\.?\s*citizen", "US Citizen"),
    (r"green\s*card", "Green Card"),
    (r"permanent\s+resident", "Permanent Resident"),
    (r"stem\s*opt", "STEM OPT"),
    (r"\bopt\b", "OPT"),
    (r"\bcpt\b", "CPT"),
    (r"h-?1b", "H-1B"),
    (r"tn\s*visa", "TN Visa"),
    (r"\bl-?1\b", "L-1"),
    (r"\be-?3\b", "E-3"),
    (r"\bf-?1\b", "F-1"),
    (r"authorized\s+to\s+work", "Authorized to work (unspecified)"),
]


def detect_visa_status(text):
    for pattern, label in VISA_PATTERNS:
        if re.search(pattern, text, re.I):
            return label
    return ""


def extract_text(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".pdf":
        parts = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                parts.append(page.extract_text() or "")
        return "\n".join(parts)
    elif ext == ".docx":
        d = docx.Document(filepath)
        return "\n".join(p.text for p in d.paragraphs)
    return ""


def parse_heuristic(text):
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    joined = "\n".join(lines)

    email = next(iter(re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", joined)), "")
    phone = next(iter(re.findall(r"(?:\+?\d{1,3}[\s.-]?)?\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}", joined)), "")
    linkedin = next(iter(re.findall(r"linkedin\.com/[a-zA-Z0-9\-_/]+", joined, re.I)), "")
    location = next(iter(re.findall(r"\b[A-Z][a-zA-Z.]+(?:\s[A-Z][a-zA-Z.]+)*,\s?[A-Z]{2}\b", joined)), "")

    name = ""
    for l in lines[:6]:
        if "@" in l or re.search(r"\d{3}", l):
            continue
        if re.match(r"^(summary|experience|education|skills|core competencies)", l, re.I):
            continue
        if 3 < len(l) < 45:
            name = l
            break

    current_title = ""
    for l in lines[:10]:
        if l == name:
            continue
        if any(k in l for k in TITLE_KEYWORDS):
            current_title = l
            break

    years_match = re.search(r"(\d+)\+?\s*years?", joined, re.I)
    years_experience = years_match.group(0) if years_match else ""

    edu_match = re.search(
        r"(Bachelor[^\n,]{0,60}|Master[^\n,]{0,60}|B\.?S\.?\s?in[^\n,]{0,50}|"
        r"M\.?S\.?\s?in[^\n,]{0,50}|B\.?A\.?\s?in[^\n,]{0,50}|Ph\.?D\.?[^\n,]{0,50}|"
        r"MBA[^\n,]{0,50}|Associate[^\n,]{0,50})", joined, re.I)
    education = edu_match.group(0).strip() if edu_match else ""

    found_skills = [s.replace("\\", "") for s in SKILLS_DICT if re.search(r"\b" + s + r"\b", joined, re.I)]
    skills = ", ".join(found_skills[:20])

    experience = []
    date_re = re.compile(
        r"((?:[A-Z][a-z]{2,8}\.?\s)?\d{4})\s?(?:-|\u2013|\u2014|to)\s?(Present|Current|(?:[A-Z][a-z]{2,8}\.?\s)?\d{4})"
    )
    used = set()
    for m in date_re.finditer(joined):
        idx = None
        for i, l in enumerate(lines):
            if i not in used and m.group(0) in l:
                idx = i
                break
        if idx is None:
            continue
        used.add(idx)
        experience.append({
            "title": lines[idx - 1] if idx > 0 else "",
            "company": "",
            "dates": m.group(0),
            "location": "",
            "highlights": (lines[idx + 1][:180] if idx + 1 < len(lines) else "")
        })
        if len(experience) >= 6:
            break

    return {
        "name": name, "email": email, "phone": phone, "linkedin": linkedin,
        "location": location, "current_title": current_title,
        "years_experience": years_experience, "education": education,
        "skills": skills, "summary": "", "visa_status": detect_visa_status(joined),
        "experience": experience
    }


def parse_with_claude(text):
    system = (
        'You are a resume parser. Read the resume text and extract structured data. '
        'Respond with ONLY a raw JSON object, no markdown fences, no commentary. Schema: '
        '{"name":"","email":"","phone":"","linkedin":"","location":"","current_title":"",'
        '"years_experience":"","education":"","skills":"comma-separated list of key skills",'
        '"summary":"1-2 sentence professional summary","visa_status":"work authorization or visa status '
        'if explicitly stated on the resume (e.g. US Citizen, Green Card, H-1B, OPT, TN Visa, '
        'requires sponsorship, no sponsorship required)",'
        '"experience":[{"title":"","company":"",'
        '"dates":"","location":"","highlights":"2-3 sentence summary of key responsibilities and achievements"}]}'
        ' Include at most the 5 most recent roles in experience, most recent first. Keep every field concise. '
        'Only fill visa_status if the resume explicitly states work authorization or visa information — '
        'never guess or infer it from nationality, name, or location. '
        'Use empty string "" for any field you cannot find. Do not invent information.'
    )
    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "Content-Type": "application/json",
            "x-api-key": ANTHROPIC_API_KEY,
            "anthropic-version": "2023-06-01",
        },
        json={
            "model": "claude-sonnet-4-6",
            "max_tokens": 1200,
            "system": system,
            "messages": [{"role": "user", "content": text[:14000]}]
        },
        timeout=60
    )
    resp.raise_for_status()
    data = resp.json()
    text_block = next((b for b in data.get("content", []) if b.get("type") == "text"), None)
    if not text_block:
        raise ValueError("No response from AI parser")
    clean = text_block["text"].strip()
    clean = re.sub(r"^```json", "", clean, flags=re.I).strip()
    clean = re.sub(r"^```", "", clean).strip()
    clean = re.sub(r"```$", "", clean).strip()
    return json.loads(clean)


# ---------------------------------------------------------------- routes
def build_candidate_filter(args):
    """Build a WHERE clause + params list from request query args for candidate search."""
    clauses = []
    params = []
    field_map = {
        "name": "name",
        "location": "location",
        "visa_status": "visa_status",
        "skills": "skills",
        "title": "current_title",
        "email": "email",
    }
    for arg_key, col in field_map.items():
        val = (args.get(arg_key) or "").strip()
        if val:
            clauses.append(f"{col} LIKE ?")
            params.append(f"%{val}%")
    where_sql = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    return where_sql, params


@app.route("/search")
@login_required
def search():
    where_sql, params = build_candidate_filter(request.args)
    conn = get_db()
    candidates = conn.execute(
        f"SELECT * FROM candidates{where_sql} ORDER BY uploaded_at DESC", params
    ).fetchall()
    conn.close()
    has_query = any((request.args.get(k) or "").strip() for k in
                     ["name", "location", "visa_status", "skills", "title", "email"])
    return render_template("search.html", candidates=candidates, args=request.args, has_query=has_query)


def parse_requirement_heuristic(text):
    # split on newlines AND sentence boundaries, since requirements are often pasted as one paragraph
    lines = [l.strip() for l in re.split(r"[\r\n]+|(?<=[.;])\s+", text) if l.strip()]
    joined = "\n".join(lines)

    def labeled(pattern):
        m = re.search(pattern, joined, re.I)
        return m.group(1).strip()[:80] if m else ""

    title = labeled(r"(?:job\s*title|title|position|role)\s*[:\-]\s*(.+)")
    location = labeled(r"location\s*[:\-]\s*(.+)")
    skills_line = labeled(r"(?:required\s*skills|skills|tech\s*stack|requirements?)\s*[:\-]\s*(.+)")
    visa_line = labeled(r"(?:visa|work\s*authorization|sponsorship)\s*[:\-]\s*(.+)")

    if not location:
        location = next(iter(re.findall(
            r"\b[A-Z][a-zA-Z.]+(?:\s[A-Z][a-zA-Z.]+)*,\s?[A-Z]{2}\b", joined)), "")

    if not title:
        for l in lines[:5]:
            if any(k in l for k in TITLE_KEYWORDS) and len(l) < 80:
                title = l
                break
        if not title and lines:
            title = lines[0][:80]

    if skills_line:
        skills = ", ".join(s.strip() for s in re.split(r",|/|;", skills_line) if s.strip())
    else:
        found = [s.replace("\\", "") for s in SKILLS_DICT if re.search(r"\b" + s + r"\b", joined, re.I)]
        skills = ", ".join(found[:20])

    visa_status = visa_line if visa_line else detect_visa_status(joined)

    return {"title": title, "location": location, "skills": skills, "visa_status": visa_status}


def parse_requirement_with_claude(text):
    system = (
        'You extract structured hiring criteria from a job requirement or job description. '
        'Respond with ONLY a raw JSON object, no markdown fences, no commentary. Schema: '
        '{"title":"job title being hired for","location":"required work location, empty if remote/unspecified",'
        '"skills":"comma-separated list of required or preferred technical skills",'
        '"visa_status":"work authorization or visa requirement if mentioned, e.g. \'US Citizen only\', '
        '\'no sponsorship\', \'H-1B ok\' — empty string if not mentioned"} '
        'Use empty string "" for any field not present. Do not invent information.'
    )
    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "Content-Type": "application/json",
            "x-api-key": ANTHROPIC_API_KEY,
            "anthropic-version": "2023-06-01",
        },
        json={
            "model": "claude-sonnet-4-6",
            "max_tokens": 500,
            "system": system,
            "messages": [{"role": "user", "content": text[:8000]}]
        },
        timeout=60
    )
    resp.raise_for_status()
    data = resp.json()
    text_block = next((b for b in data.get("content", []) if b.get("type") == "text"), None)
    if not text_block:
        raise ValueError("No response from AI parser")
    clean = text_block["text"].strip()
    clean = re.sub(r"^```json", "", clean, flags=re.I).strip()
    clean = re.sub(r"^```", "", clean).strip()
    clean = re.sub(r"```$", "", clean).strip()
    return json.loads(clean)


def score_candidates_against_requirement(req_row, candidates):
    req_skills = [s.strip().lower() for s in (req_row["skills"] or "").split(",") if s.strip()]
    req_location = (req_row["location"] or "").strip().lower()
    req_visa = (req_row["visa_status"] or "").strip().lower()

    scored = []
    for c in candidates:
        if req_location and req_location not in (c["location"] or "").lower():
            continue
        if req_visa and req_visa not in (c["visa_status"] or "").lower():
            continue
        cand_skills = [s.strip().lower() for s in (c["skills"] or "").split(",") if s.strip()]
        overlap = sorted(set(req_skills) & set(cand_skills))
        if req_skills and not overlap:
            continue
        scored.append({"candidate": c, "match_count": len(overlap), "matched_skills": overlap})

    scored.sort(key=lambda r: r["match_count"], reverse=True)
    return scored


@app.route("/requirements", methods=["GET", "POST"])
@login_required
def requirements():
    conn = get_db()
    if request.method == "POST":
        raw_text = request.form.get("raw_text", "").strip()
        if not raw_text:
            flash("Paste a job requirement first.")
            conn.close()
            return redirect(url_for("requirements"))

        if ANTHROPIC_API_KEY:
            try:
                parsed = parse_requirement_with_claude(raw_text)
            except Exception:
                parsed = parse_requirement_heuristic(raw_text)
        else:
            parsed = parse_requirement_heuristic(raw_text)

        label = (parsed.get("title") or raw_text[:60]).strip()
        cur = conn.execute(
            """INSERT INTO requirements (label, raw_text, title, location, skills, visa_status,
               created_by, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (label, raw_text, parsed.get("title", ""), parsed.get("location", ""),
             parsed.get("skills", ""), parsed.get("visa_status", ""),
             session.get("username", ""), datetime.utcnow().isoformat())
        )
        conn.commit()
        req_id = cur.lastrowid
        conn.close()
        return redirect(url_for("view_requirement", req_id=req_id))

    saved = conn.execute("SELECT * FROM requirements ORDER BY created_at DESC").fetchall()
    conn.close()
    return render_template("requirements.html", saved=saved, current=None, matches=None,
                            ai_enabled=bool(ANTHROPIC_API_KEY))


@app.route("/requirements/<int:req_id>")
@login_required
def view_requirement(req_id):
    conn = get_db()
    current = conn.execute("SELECT * FROM requirements WHERE id = ?", (req_id,)).fetchone()
    if not current:
        conn.close()
        flash("That saved requirement no longer exists.")
        return redirect(url_for("requirements"))
    candidates = conn.execute("SELECT * FROM candidates ORDER BY uploaded_at DESC").fetchall()
    saved = conn.execute("SELECT * FROM requirements ORDER BY created_at DESC").fetchall()
    conn.close()
    matches = score_candidates_against_requirement(current, candidates)
    return render_template("requirements.html", saved=saved, current=current, matches=matches,
                            ai_enabled=bool(ANTHROPIC_API_KEY))


@app.route("/requirements/<int:req_id>/delete", methods=["POST"])
@login_required
def delete_requirement(req_id):
    conn = get_db()
    conn.execute("DELETE FROM requirements WHERE id = ?", (req_id,))
    conn.commit()
    conn.close()
    return redirect(url_for("requirements"))


@app.route("/")
@login_required
def dashboard():
    conn = get_db()
    candidates = conn.execute("SELECT * FROM candidates ORDER BY uploaded_at DESC").fetchall()
    users = conn.execute("SELECT username FROM users ORDER BY username").fetchall()
    conn.close()
    return render_template("dashboard.html", candidates=candidates, users=users,
                            ai_enabled=bool(ANTHROPIC_API_KEY))


@app.route("/upload", methods=["POST"])
@login_required
def upload():
    files = request.files.getlist("resumes")
    conn = get_db()
    results = []
    for f in files:
        if not f or not f.filename:
            continue
        save_path = None
        try:
            ext = os.path.splitext(f.filename)[1].lower()
            if ext not in ALLOWED_EXT:
                results.append({"filename": f.filename, "status": "skipped", "reason": "unsupported file type"})
                continue
            safe_name = secure_filename(f.filename)
            save_path = os.path.join(UPLOAD_DIR, f"{datetime.utcnow().timestamp()}_{safe_name}")
            f.save(save_path)

            text = extract_text(save_path)
            if not text or len(text.strip()) < 20:
                raise ValueError("No readable text found (possibly a scanned/image-only file)")

            if ANTHROPIC_API_KEY:
                try:
                    parsed = parse_with_claude(text)
                except Exception:
                    parsed = parse_heuristic(text)
            else:
                parsed = parse_heuristic(text)

            cur = conn.execute(
                """INSERT INTO candidates
                (filename, name, email, phone, linkedin, location, current_title,
                 years_experience, education, skills, summary, visa_status, uploaded_by, uploaded_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (f.filename, parsed.get("name", ""), parsed.get("email", ""), parsed.get("phone", ""),
                 parsed.get("linkedin", ""), parsed.get("location", ""), parsed.get("current_title", ""),
                 parsed.get("years_experience", ""), parsed.get("education", ""), parsed.get("skills", ""),
                 parsed.get("summary", ""), parsed.get("visa_status", ""),
                 session.get("username", ""), datetime.utcnow().isoformat())
            )
            candidate_id = cur.lastrowid
            for exp in parsed.get("experience", []):
                conn.execute(
                    """INSERT INTO experience (candidate_id, title, company, dates, location, highlights)
                    VALUES (?, ?, ?, ?, ?, ?)""",
                    (candidate_id, exp.get("title", ""), exp.get("company", ""), exp.get("dates", ""),
                     exp.get("location", ""), exp.get("highlights", ""))
                )
            conn.commit()
            results.append({"filename": f.filename, "status": "done"})
        except Exception as e:
            conn.rollback()
            results.append({"filename": f.filename, "status": "error", "reason": str(e)[:200]})
        finally:
            if save_path:
                try:
                    os.remove(save_path)
                except OSError:
                    pass
    conn.close()
    return jsonify({"results": results})


@app.errorhandler(500)
def handle_500(e):
    if request.path.startswith("/upload"):
        return jsonify({"results": [{"filename": "", "status": "error",
                                      "reason": "Server error — check Render logs for details"}]}), 500
    return e


@app.errorhandler(413)
def handle_413(e):
    if request.path.startswith("/upload"):
        return jsonify({"results": [{"filename": "", "status": "error",
                                      "reason": "File too large (25MB limit per request)"}]}), 413
    return e


@app.route("/candidate/<int:candidate_id>/delete", methods=["POST"])
@login_required
def delete_candidate(candidate_id):
    conn = get_db()
    conn.execute("DELETE FROM experience WHERE candidate_id = ?", (candidate_id,))
    conn.execute("DELETE FROM candidates WHERE id = ?", (candidate_id,))
    conn.commit()
    conn.close()
    return redirect(url_for("dashboard"))


@app.route("/export")
@login_required
def export():
    where_sql, params = build_candidate_filter(request.args)
    conn = get_db()
    candidates = conn.execute(
        f"SELECT * FROM candidates{where_sql} ORDER BY uploaded_at DESC", params
    ).fetchall()
    candidate_ids = [c["id"] for c in candidates]
    if candidate_ids:
        placeholders = ",".join("?" * len(candidate_ids))
        experiences = conn.execute(
            f"SELECT * FROM experience WHERE candidate_id IN ({placeholders})", candidate_ids
        ).fetchall()
    else:
        experiences = []
    conn.close()

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Candidates"
    headers1 = ["File", "Name", "Current Title", "Email", "Phone", "LinkedIn", "Location",
                "Years Experience", "Education", "Skills", "Visa Status", "Summary",
                "Uploaded By", "Uploaded At"]
    ws1.append(headers1)
    for c in candidates:
        ws1.append([c["filename"], c["name"], c["current_title"], c["email"], c["phone"],
                    c["linkedin"], c["location"], c["years_experience"], c["education"],
                    c["skills"], c["visa_status"], c["summary"], c["uploaded_by"], c["uploaded_at"]])
    widths1 = [24, 20, 22, 26, 16, 26, 18, 14, 30, 40, 20, 45, 14, 20]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Work Experience")
    cand_map = {c["id"]: c["name"] for c in candidates}
    headers2 = ["Candidate", "Job Title", "Company", "Dates", "Location", "Highlights"]
    ws2.append(headers2)
    for e in experiences:
        ws2.append([cand_map.get(e["candidate_id"], ""), e["title"], e["company"], e["dates"],
                    e["location"], e["highlights"]])
    widths2 = [20, 22, 20, 16, 18, 55]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="Resumes_Parsed.xlsx",
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


init_db()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
